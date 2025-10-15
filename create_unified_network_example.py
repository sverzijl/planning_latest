"""Create example Network_Config_Unified.xlsx in unified model format.

This demonstrates the new unified format with:
- Nodes sheet (with capability flags)
- Routes with origin_node_id
- TruckSchedules with origin_node_id (enables hub trucks!)
"""

import pandas as pd
from datetime import date, time, timedelta
import openpyxl
from openpyxl import Workbook

# Create new workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

print("Creating Network_Config_Unified.xlsx...")
print("=" * 80)

# ==================
# SHEET 1: NODES
# ==================

print("\n1. Creating Nodes sheet...")

nodes_data = [
    # Manufacturing node
    {
        'node_id': '6122',
        'node_name': 'Manufacturing Site',
        'can_manufacture': True,
        'production_rate_per_hour': 1400.0,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': None,
        'has_demand': False,
        'requires_truck_schedules': True,
        'latitude': -37.7,
        'longitude': 144.9,
    },
    # Hubs with demand
    {
        'node_id': '6104',
        'node_name': 'NSW Hub (Moorebank)',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': 50000,
        'has_demand': True,
        'requires_truck_schedules': False,  # Could be True to add hub truck constraints
        'latitude': -33.9,
        'longitude': 150.9,
    },
    {
        'node_id': '6125',
        'node_name': 'VIC Hub (Keilor Park)',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': 50000,
        'has_demand': True,
        'requires_truck_schedules': False,
        'latitude': -37.7,
        'longitude': 144.8,
    },
    # Frozen storage (intermediate)
    {
        'node_id': 'Lineage',
        'node_name': 'Lineage Frozen Storage',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'frozen',  # Frozen only
        'storage_capacity': 100000,
        'has_demand': False,
        'requires_truck_schedules': False,
        'latitude': -31.9,
        'longitude': 115.9,
    },
    # Spoke locations
    {
        'node_id': '6103',
        'node_name': 'Canberra',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': None,
        'has_demand': True,
        'requires_truck_schedules': False,
        'latitude': -35.3,
        'longitude': 149.1,
    },
    {
        'node_id': '6105',
        'node_name': 'Rydalmere (NSW)',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': None,
        'has_demand': True,
        'requires_truck_schedules': False,
        'latitude': -33.8,
        'longitude': 151.0,
    },
    {
        'node_id': '6110',
        'node_name': 'Queensland (Direct)',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': None,
        'has_demand': True,
        'requires_truck_schedules': False,
        'latitude': -27.9,
        'longitude': 153.4,
    },
    {
        'node_id': '6123',
        'node_name': 'Clayton (VIC)',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': None,
        'has_demand': True,
        'requires_truck_schedules': False,
        'latitude': -37.9,
        'longitude': 145.1,
    },
    {
        'node_id': '6120',
        'node_name': 'Hobart (TAS)',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'ambient',
        'storage_capacity': None,
        'has_demand': True,
        'requires_truck_schedules': False,
        'latitude': -42.9,
        'longitude': 147.3,
    },
    # WA with freeze/thaw capability
    {
        'node_id': '6130',
        'node_name': 'WA Breadroom (Thawing)',
        'can_manufacture': False,
        'production_rate_per_hour': None,
        'can_store': True,
        'storage_mode': 'both',  # Can handle frozen AND ambient (thaws on-site)
        'storage_capacity': None,
        'has_demand': True,
        'requires_truck_schedules': False,
        'latitude': -32.0,
        'longitude': 115.9,
    },
]

df_nodes = pd.DataFrame(nodes_data)
ws_nodes = wb.create_sheet("Nodes", 0)

# Write header
for col_idx, col_name in enumerate(df_nodes.columns, start=1):
    ws_nodes.cell(row=1, column=col_idx, value=col_name)

# Write data
for row_idx, row_data in enumerate(df_nodes.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_nodes.cell(row=row_idx, column=col_idx, value=value)

print(f"   {len(nodes_data)} nodes created")

# ==================
# SHEET 2: ROUTES
# ==================

print("\n2. Creating Routes sheet...")

routes_data = [
    # Manufacturing to hubs
    {'route_id': 'R1', 'origin_node_id': '6122', 'destination_node_id': '6104', 'transit_days': 1.0, 'transport_mode': 'ambient', 'cost_per_unit': 0.30},
    {'route_id': 'R2', 'origin_node_id': '6122', 'destination_node_id': '6110', 'transit_days': 1.5, 'transport_mode': 'ambient', 'cost_per_unit': 0.40},
    {'route_id': 'R3', 'origin_node_id': '6122', 'destination_node_id': '6125', 'transit_days': 1.0, 'transport_mode': 'ambient', 'cost_per_unit': 0.30},
    {'route_id': 'R4', 'origin_node_id': '6122', 'destination_node_id': 'Lineage', 'transit_days': 0.5, 'transport_mode': 'ambient', 'cost_per_unit': 0.20},

    # Hub to spokes (NSW)
    {'route_id': 'R5', 'origin_node_id': '6104', 'destination_node_id': '6105', 'transit_days': 0.5, 'transport_mode': 'ambient', 'cost_per_unit': 0.15},
    {'route_id': 'R6', 'origin_node_id': '6104', 'destination_node_id': '6103', 'transit_days': 1.0, 'transport_mode': 'ambient', 'cost_per_unit': 0.20},

    # Hub to spokes (VIC)
    {'route_id': 'R7', 'origin_node_id': '6125', 'destination_node_id': '6123', 'transit_days': 0.5, 'transport_mode': 'ambient', 'cost_per_unit': 0.15},
    {'route_id': 'R8', 'origin_node_id': '6125', 'destination_node_id': '6120', 'transit_days': 2.0, 'transport_mode': 'ambient', 'cost_per_unit': 0.25},

    # Frozen route (Lineage to WA)
    {'route_id': 'R10', 'origin_node_id': 'Lineage', 'destination_node_id': '6130', 'transit_days': 3.0, 'transport_mode': 'frozen', 'cost_per_unit': 0.50},
]

df_routes = pd.DataFrame(routes_data)
ws_routes = wb.create_sheet("Routes", 1)

for col_idx, col_name in enumerate(df_routes.columns, start=1):
    ws_routes.cell(row=1, column=col_idx, value=col_name)

for row_idx, row_data in enumerate(df_routes.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_routes.cell(row=row_idx, column=col_idx, value=value)

print(f"   {len(routes_data)} routes created")

# ==================
# SHEET 3: TRUCK SCHEDULES
# ==================

print("\n3. Creating TruckSchedules sheet...")

trucks_data = [
    # Manufacturing trucks (Mon-Fri to hubs)
    {'truck_id': 'T1', 'truck_name': 'Morning 6125 (Mon)', 'origin_node_id': '6122', 'destination_node_id': '6125', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'monday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.30, 'intermediate_stops': None},
    {'truck_id': 'T2', 'truck_name': 'Morning 6125 (Tue)', 'origin_node_id': '6122', 'destination_node_id': '6125', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'tuesday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.30, 'intermediate_stops': None},
    {'truck_id': 'T3', 'truck_name': 'Morning 6125 via Lineage (Wed)', 'origin_node_id': '6122', 'destination_node_id': '6125', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'wednesday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.35, 'intermediate_stops': 'Lineage'},
    {'truck_id': 'T4', 'truck_name': 'Morning 6125 (Thu)', 'origin_node_id': '6122', 'destination_node_id': '6125', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'thursday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.30, 'intermediate_stops': None},
    {'truck_id': 'T5', 'truck_name': 'Morning 6125 (Fri)', 'origin_node_id': '6122', 'destination_node_id': '6125', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'friday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.30, 'intermediate_stops': None},

    # Manufacturing afternoon trucks
    {'truck_id': 'T6', 'truck_name': 'Afternoon 6104 (Mon)', 'origin_node_id': '6122', 'destination_node_id': '6104', 'departure_type': 'afternoon', 'departure_time': '14:00', 'day_of_week': 'monday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.30, 'intermediate_stops': None},
    {'truck_id': 'T7', 'truck_name': 'Afternoon 6110 (Tue)', 'origin_node_id': '6122', 'destination_node_id': '6110', 'departure_type': 'afternoon', 'departure_time': '14:00', 'day_of_week': 'tuesday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.40, 'intermediate_stops': None},
    {'truck_id': 'T8', 'truck_name': 'Afternoon 6104 (Wed)', 'origin_node_id': '6122', 'destination_node_id': '6104', 'departure_type': 'afternoon', 'departure_time': '14:00', 'day_of_week': 'wednesday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.30, 'intermediate_stops': None},
    {'truck_id': 'T9', 'truck_name': 'Afternoon 6110 (Thu)', 'origin_node_id': '6122', 'destination_node_id': '6110', 'departure_type': 'afternoon', 'departure_time': '14:00', 'day_of_week': 'thursday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.40, 'intermediate_stops': None},
    {'truck_id': 'T10', 'truck_name': 'Afternoon 6110 (Fri)', 'origin_node_id': '6122', 'destination_node_id': '6110', 'departure_type': 'afternoon', 'departure_time': '14:00', 'day_of_week': 'friday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.40, 'intermediate_stops': None},
    {'truck_id': 'T11', 'truck_name': 'Afternoon 6104 (Fri #2)', 'origin_node_id': '6122', 'destination_node_id': '6104', 'departure_type': 'afternoon', 'departure_time': '14:00', 'day_of_week': 'friday', 'capacity': 14080, 'cost_fixed': 100.0, 'cost_per_unit': 0.30, 'intermediate_stops': None},

    # HUB TRUCKS (NEW CAPABILITY!)
    # These trucks originate from hubs - NOT possible in legacy format!
    {'truck_id': 'T_HUB_VIC_1', 'truck_name': 'VIC Hub → Clayton (Mon/Wed/Fri)', 'origin_node_id': '6125', 'destination_node_id': '6123', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'monday', 'capacity': 14080, 'cost_fixed': 80.0, 'cost_per_unit': 0.15, 'intermediate_stops': None},
    {'truck_id': 'T_HUB_VIC_2', 'truck_name': 'VIC Hub → Clayton (Wed)', 'origin_node_id': '6125', 'destination_node_id': '6123', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'wednesday', 'capacity': 14080, 'cost_fixed': 80.0, 'cost_per_unit': 0.15, 'intermediate_stops': None},
    {'truck_id': 'T_HUB_VIC_3', 'truck_name': 'VIC Hub → Clayton (Fri)', 'origin_node_id': '6125', 'destination_node_id': '6123', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'friday', 'capacity': 14080, 'cost_fixed': 80.0, 'cost_per_unit': 0.15, 'intermediate_stops': None},

    {'truck_id': 'T_HUB_NSW_1', 'truck_name': 'NSW Hub → Rydalmere (Tue/Thu)', 'origin_node_id': '6104', 'destination_node_id': '6105', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'tuesday', 'capacity': 14080, 'cost_fixed': 80.0, 'cost_per_unit': 0.15, 'intermediate_stops': None},
    {'truck_id': 'T_HUB_NSW_2', 'truck_name': 'NSW Hub → Rydalmere (Thu)', 'origin_node_id': '6104', 'destination_node_id': '6105', 'departure_type': 'morning', 'departure_time': '08:00', 'day_of_week': 'thursday', 'capacity': 14080, 'cost_fixed': 80.0, 'cost_per_unit': 0.15, 'intermediate_stops': None},
]

df_trucks = pd.DataFrame(trucks_data)
ws_trucks = wb.create_sheet("TruckSchedules", 2)

for col_idx, col_name in enumerate(df_trucks.columns, start=1):
    ws_trucks.cell(row=1, column=col_idx, value=col_name)

for row_idx, row_data in enumerate(df_trucks.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_trucks.cell(row=row_idx, column=col_idx, value=value)

print(f"   {len(trucks_data)} truck schedules created")
print(f"   - {sum(1 for t in trucks_data if t['origin_node_id'] == '6122')} from manufacturing")
print(f"   - {sum(1 for t in trucks_data if t['origin_node_id'] != '6122')} from hubs (NEW!)")

# ==================
# SHEET 4: COST PARAMETERS
# ==================

print("\n4. Creating CostParameters sheet...")

costs_data = [
    {'cost_type': 'production_cost_per_unit', 'value': 5.00, 'unit': '$/unit'},
    {'cost_type': 'transport_cost_per_unit', 'value': 1.00, 'unit': '$/unit'},
    {'cost_type': 'storage_cost_per_unit_per_day', 'value': 0.01, 'unit': '$/unit/day'},
    {'cost_type': 'waste_cost_per_unit', 'value': 10.00, 'unit': '$/unit'},
    {'cost_type': 'shortage_penalty_per_unit', 'value': 10000.00, 'unit': '$/unit'},  # High penalty!
]

df_costs = pd.DataFrame(costs_data)
ws_costs = wb.create_sheet("CostParameters", 3)

for col_idx, col_name in enumerate(df_costs.columns, start=1):
    ws_costs.cell(row=1, column=col_idx, value=col_name)

for row_idx, row_data in enumerate(df_costs.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_costs.cell(row=row_idx, column=col_idx, value=value)

print(f"   {len(costs_data)} cost parameters created")

# ==================
# SHEET 5: LABOR CALENDAR (Sample)
# ==================

print("\n5. Creating sample LaborCalendar sheet...")

# Create 30 days of sample labor data
labor_data = []
start_date = date(2025, 10, 1)

for i in range(30):
    current_date = start_date + timedelta(days=i)
    weekday = current_date.weekday()

    if weekday < 5:  # Monday-Friday
        labor_data.append({
            'date': current_date,
            'fixed_hours': 12.0,
            'regular_rate': 25.0,
            'overtime_rate': 37.5,
            'non_fixed_rate': None,
            'minimum_hours': 0.0,
            'is_fixed_day': True,
        })
    else:  # Weekend
        labor_data.append({
            'date': current_date,
            'fixed_hours': 0.0,
            'regular_rate': 25.0,
            'overtime_rate': 37.5,
            'non_fixed_rate': 50.0,
            'minimum_hours': 4.0,
            'is_fixed_day': False,
        })

df_labor = pd.DataFrame(labor_data)
ws_labor = wb.create_sheet("LaborCalendar", 4)

for col_idx, col_name in enumerate(df_labor.columns, start=1):
    ws_labor.cell(row=1, column=col_idx, value=col_name)

for row_idx, row_data in enumerate(df_labor.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_labor.cell(row=row_idx, column=col_idx, value=value)

print(f"   {len(labor_data)} labor days created")

# Save workbook
output_file = 'data/examples/Network_Config_Unified.xlsx'
wb.save(output_file)

print(f"\n✅ Created: {output_file}")
print("\nKEY FEATURES DEMONSTRATED:")
print("  ✓ Nodes with capability flags (can_manufacture, has_demand, etc.)")
print("  ✓ storage_mode options: frozen, ambient, both")
print("  ✓ Routes with explicit origin_node_id and destination_node_id")
print("  ✓ TruckSchedules with origin_node_id (enables hub trucks!)")
print("  ✓ Hub truck schedules: 6125→6123, 6104→6105")
print("\nThis format enables features not possible in legacy format!")
