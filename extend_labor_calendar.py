#!/usr/bin/env python3
"""
Extend Labor Calendar in Network_Config.xlsx from Dec 23, 2025 through Dec 31, 2026.

This script:
1. Backs up the original file
2. Loads the LaborCalendar sheet
3. Generates 374 days of labor calendar entries
4. Applies correct patterns based on weekday/weekend/holiday
5. Saves the updated file
6. Validates the results
"""

import openpyxl
from datetime import date, timedelta
import shutil
from pathlib import Path

# File paths
EXCEL_FILE = Path('/home/sverzijl/planning_latest/data/examples/Network_Config.xlsx')
BACKUP_FILE = Path('/home/sverzijl/planning_latest/data/examples/Network_Config_backup.xlsx')

# Victoria Public Holidays (2025 end-of-year + all 2026)
PUBLIC_HOLIDAYS = {
    # 2025 (end of year - in our extension range)
    date(2025, 12, 25), # Christmas Day (Thursday)
    date(2025, 12, 26), # Boxing Day (Friday)
    date(2025, 12, 29), # Boxing Day substitute (Monday) - since Boxing Day falls on Friday

    # 2026
    date(2026, 1, 1),   # New Year's Day (Thursday)
    date(2026, 1, 26),  # Australia Day (Monday)
    date(2026, 3, 9),   # Labour Day (Monday)
    date(2026, 4, 3),   # Good Friday
    date(2026, 4, 4),   # Saturday before Easter Sunday
    date(2026, 4, 5),   # Easter Sunday
    date(2026, 4, 6),   # Easter Monday
    date(2026, 4, 25),  # ANZAC Day (Saturday - no substitute)
    date(2026, 6, 8),   # King's Birthday (Monday)
    date(2026, 9, 25),  # AFL Grand Final Friday (estimated)
    date(2026, 11, 3),  # Melbourne Cup Day (Tuesday)
    date(2026, 12, 25), # Christmas Day (Friday)
    date(2026, 12, 26), # Boxing Day (Saturday)
    date(2026, 12, 28), # Boxing Day substitute (Monday)
}

# Labor patterns
WEEKDAY_PATTERN = {
    'fixed_hours': 12,
    'regular_rate': 25.00,
    'overtime_rate': 37.50,
    'non_fixed_rate': None,
    'minimum_hours': 0,
    'is_fixed_day': True
}

WEEKEND_HOLIDAY_PATTERN = {
    'fixed_hours': 0,
    'regular_rate': 25.00,
    'overtime_rate': 37.50,
    'non_fixed_rate': 40.00,
    'minimum_hours': 4,
    'is_fixed_day': False
}

def get_labor_pattern(target_date: date) -> dict:
    """Determine labor pattern based on date."""
    # Check if public holiday
    if target_date in PUBLIC_HOLIDAYS:
        return WEEKEND_HOLIDAY_PATTERN.copy()

    # Check if weekend (0=Monday, 6=Sunday)
    weekday = target_date.weekday()
    if weekday >= 5:  # Saturday=5, Sunday=6
        return WEEKEND_HOLIDAY_PATTERN.copy()

    # Regular weekday
    return WEEKDAY_PATTERN.copy()

def main():
    print("="*80)
    print("LABOR CALENDAR EXTENSION SCRIPT")
    print("="*80)
    print()

    # 1. Backup file
    print(f"1. Creating backup: {BACKUP_FILE}")
    shutil.copy(EXCEL_FILE, BACKUP_FILE)
    print(f"   ✓ Backup created")
    print()

    # 2. Load workbook
    print(f"2. Loading workbook: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['LaborCalendar']
    print(f"   ✓ Current rows: {ws.max_row}")

    # Get last date in current calendar
    last_row = ws.max_row
    last_date_cell = ws.cell(row=last_row, column=1)
    last_date = last_date_cell.value.date() if hasattr(last_date_cell.value, 'date') else last_date_cell.value
    print(f"   ✓ Last date in calendar: {last_date}")
    print()

    # 3. Generate new rows
    start_date = date(2025, 12, 23)
    end_date = date(2026, 12, 31)

    print(f"3. Generating rows from {start_date} to {end_date}")

    current_date = start_date
    rows_added = 0
    weekday_count = 0
    weekend_count = 0
    holiday_count = 0

    while current_date <= end_date:
        # Get labor pattern
        pattern = get_labor_pattern(current_date)

        # Count categories
        if current_date in PUBLIC_HOLIDAYS:
            holiday_count += 1
        elif current_date.weekday() >= 5:
            weekend_count += 1
        else:
            weekday_count += 1

        # Append row to sheet
        ws.append([
            current_date,
            pattern['fixed_hours'],
            pattern['regular_rate'],
            pattern['overtime_rate'],
            pattern['non_fixed_rate'],
            pattern['minimum_hours'],
            pattern['is_fixed_day']
        ])

        rows_added += 1
        current_date += timedelta(days=1)

    print(f"   ✓ Rows added: {rows_added}")
    print(f"   ✓ Weekdays: {weekday_count}")
    print(f"   ✓ Weekend days: {weekend_count}")
    print(f"   ✓ Public holidays: {holiday_count}")
    print()

    # 4. Save workbook
    print("4. Saving updated workbook")
    wb.save(EXCEL_FILE)
    print(f"   ✓ Saved to {EXCEL_FILE}")
    print()

    # 5. Validate
    print("5. Validation")
    wb_validate = openpyxl.load_workbook(EXCEL_FILE)
    ws_validate = wb_validate['LaborCalendar']

    new_total_rows = ws_validate.max_row
    print(f"   ✓ Total rows in sheet: {new_total_rows} (was {last_row})")
    print(f"   ✓ Rows added: {new_total_rows - last_row}")

    # Check first and last new entries
    first_new_row = last_row + 1
    first_new_date = ws_validate.cell(row=first_new_row, column=1).value
    last_new_date = ws_validate.cell(row=new_total_rows, column=1).value

    print(f"   ✓ First new entry: {first_new_date}")
    print(f"   ✓ Last new entry: {last_new_date}")
    print()

    # Summary statistics
    print("="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Total rows added: {rows_added}")
    print(f"Expected rows: 374")
    print(f"Match: {'✓ YES' if rows_added == 374 else '✗ NO'}")
    print()
    print(f"Breakdown:")
    print(f"  Weekdays (fixed_hours=12): {weekday_count}")
    print(f"  Weekend days: {weekend_count}")
    print(f"  Public holidays: {holiday_count}")
    print(f"  Total non-fixed days: {weekend_count + holiday_count}")
    print()
    print(f"Date range: {start_date} to {end_date}")
    print()
    print("Public Holidays in extension period:")
    for holiday in sorted(PUBLIC_HOLIDAYS):
        if start_date <= holiday <= end_date:
            weekday_name = holiday.strftime('%A')
            print(f"  {holiday} ({weekday_name})")
    print()
    print("✓ Labor calendar successfully extended through Dec 31, 2026")
    print("="*80)

if __name__ == '__main__':
    main()
