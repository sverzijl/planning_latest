#!/usr/bin/env python3
"""
Fix Products sheet to include all 5 products with correct IDs from Alias sheet.

Root cause identified: Products sheet only has 2 products (G144, G610) with wrong IDs.
Should have 5 products with Alias1 column values as product_id.
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def fix_products_sheet(config_file: Path):
    """Fix Products sheet with all 5 products and correct IDs."""

    print(f"Fixing: {config_file}")

    # Read Alias sheet to get correct product IDs
    df_alias = pd.read_excel(config_file, sheet_name='Alias', engine='openpyxl')

    print(f"\nAlias sheet has {len(df_alias)} products:")
    for idx, row in df_alias.iterrows():
        print(f"  {row['Alias1']}")

    # Create Products sheet with correct data
    products_data = []
    mix_sizes = [415, 387, 520, 450, 395]  # From design document

    for idx, row in df_alias.iterrows():
        product_id = row['Alias1']  # Use canonical name from Alias1
        products_data.append({
            'product_id': product_id,
            'name': product_id,  # Use same as ID for simplicity
            'sku': product_id,
            'shelf_life_ambient_days': 17,
            'shelf_life_frozen_days': 120,
            'shelf_life_after_thaw_days': 14,
            'min_acceptable_shelf_life_days': 7,
            'units_per_mix': mix_sizes[idx]
        })

    df_products = pd.DataFrame(products_data)

    print(f"\nCreated Products dataframe:")
    print(df_products[['product_id', 'units_per_mix']].to_string())

    # Write to Excel with formatting
    wb = load_workbook(config_file)

    # Remove old Products sheet if exists
    if 'Products' in wb.sheetnames:
        del wb['Products']

    # Create new Products sheet as first sheet
    ws = wb.create_sheet('Products', 0)

    # Write headers
    headers = list(df_products.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, row_data in enumerate(df_products.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Set row height for header
    ws.row_dimensions[1].height = 25

    # Save
    wb.save(config_file)
    print(f"\n✓ Saved: {config_file}")

    # Verify
    df_verify = pd.read_excel(config_file, sheet_name='Products', engine='openpyxl')
    print(f"\n✓ Verification: Products sheet now has {len(df_verify)} products")

    return df_verify

# Fix all 3 config files
config_files = [
    Path("data/examples/Network_Config.xlsx"),
    Path("data/examples/Network_Config_Unified.xlsx"),
    Path("data/examples/Network_Config_backup.xlsx")
]

print("="*80)
print("FIXING PRODUCTS SHEET IN ALL NETWORK CONFIG FILES")
print("="*80)

for config_file in config_files:
    if config_file.exists():
        fix_products_sheet(config_file)
        print()
    else:
        print(f"⚠️  Skipping {config_file} (not found)\n")

print("="*80)
print("FIX COMPLETE")
print("="*80)
