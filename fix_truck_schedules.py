"""Fix Network_Config.xlsx to add missing day_of_week values."""

import openpyxl
from pathlib import Path

def fix_truck_schedules():
    """Add missing day_of_week values to truck schedules."""

    file_path = Path("data/examples/Network_Config.xlsx")

    print("=" * 80)
    print("FIXING TRUCK SCHEDULES IN NETWORK_CONFIG.XLSX")
    print("=" * 80)
    print()

    # Load workbook
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)

    # Get TruckSchedules sheet
    if 'TruckSchedules' not in wb.sheetnames:
        print("ERROR: TruckSchedules sheet not found!")
        return

    ws = wb['TruckSchedules']

    # Find column headers
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

    print(f"Found columns: {', '.join(headers.keys())}")
    print()

    if 'id' not in headers or 'day_of_week' not in headers:
        print("ERROR: Required columns not found!")
        return

    id_col = headers['id']
    day_col = headers['day_of_week']

    # Define the fixes needed
    fixes = {
        'T1': 'monday',
        'T2': 'tuesday',
        'T4': 'thursday',
        'T5': 'friday',
    }

    print("BEFORE FIXES:")
    print("-" * 80)

    # First pass: show current state
    for row_idx in range(2, ws.max_row + 1):
        truck_id = ws.cell(row=row_idx, column=id_col).value
        if truck_id in fixes:
            current_value = ws.cell(row=row_idx, column=day_col).value
            print(f"Row {row_idx}: {truck_id} - day_of_week = {current_value if current_value else 'NULL'}")

    print()
    print("APPLYING FIXES:")
    print("-" * 80)

    # Second pass: apply fixes
    fixes_applied = []
    for row_idx in range(2, ws.max_row + 1):
        truck_id = ws.cell(row=row_idx, column=id_col).value
        if truck_id in fixes:
            old_value = ws.cell(row=row_idx, column=day_col).value
            new_value = fixes[truck_id]
            ws.cell(row=row_idx, column=day_col, value=new_value)
            print(f"Row {row_idx}: {truck_id} - '{old_value if old_value else 'NULL'}' → '{new_value}'")
            fixes_applied.append(truck_id)

    print()

    if not fixes_applied:
        print("WARNING: No trucks found to fix!")
        return

    # Save the workbook
    print(f"Saving changes to {file_path}...")
    wb.save(file_path)
    print("✅ File saved successfully!")
    print()

    print("SUMMARY:")
    print("-" * 80)
    print(f"Fixed {len(fixes_applied)} truck schedules:")
    for truck_id in fixes_applied:
        print(f"  - {truck_id}: now runs on {fixes[truck_id]} only")
    print()
    print("Next steps:")
    print("  1. Verify the changes by running: venv/bin/python check_lineage_routing.py")
    print("  2. Re-run optimization to see correct truck scheduling")
    print()


if __name__ == "__main__":
    fix_truck_schedules()
