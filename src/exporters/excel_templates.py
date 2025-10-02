"""
Excel export templates for production planning results.

This module provides formatted Excel exports for:
1. Production schedules (manufacturing template)
2. Cost breakdowns (management template)
3. Shipment plans (logistics template)

All exports include professional formatting, charts, and aggregations.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Optional, List
import io

# Color constants (matching design system)
PRIMARY_COLOR = "1E88E5"
SECONDARY_COLOR = "43A047"
ACCENT_COLOR = "FB8C00"
ERROR_COLOR = "E53935"
WARNING_COLOR = "FBC02D"
HEADER_COLOR = "1E88E5"
ALT_ROW_COLOR = "F5F5F5"
HIGH_UTIL_COLOR = "C8E6C9"  # Green
MEDIUM_UTIL_COLOR = "BBDEFB"  # Blue
LOW_UTIL_COLOR = "FFF9C4"  # Yellow
OVERLOAD_COLOR = "FFCDD2"  # Red


def create_header_style() -> Dict[str, Any]:
    """Create header row style (blue background, white text, bold)."""
    return {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_cell_style(bold: bool = False, bg_color: Optional[str] = None) -> Dict[str, Any]:
    """Create standard cell style."""
    style = {
        'font': Font(name='Calibri', size=10, bold=bold),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    if bg_color:
        style['fill'] = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    return style


def apply_alternating_rows(worksheet, start_row: int, end_row: int, start_col: int = 1, end_col: int = 10):
    """Apply alternating row colors (white / light gray)."""
    for row_idx in range(start_row, end_row + 1):
        if (row_idx - start_row) % 2 == 1:  # Alternate rows
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type='solid')


def add_filters(worksheet, end_column: int, header_row: int = 1):
    """Add Excel filters to header row."""
    end_col_letter = get_column_letter(end_column)
    worksheet.auto_filter.ref = f"A{header_row}:{end_col_letter}{header_row}"


def format_currency(worksheet, column: int, start_row: int, end_row: int):
    """Format column as currency ($#,##0.00)."""
    for row_idx in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row_idx, column=column)
        cell.number_format = '$#,##0.00'


def format_number(worksheet, column: int, start_row: int, end_row: int):
    """Format column as number with thousands separator (#,##0)."""
    for row_idx in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row_idx, column=column)
        cell.number_format = '#,##0'


def format_percentage(worksheet, column: int, start_row: int, end_row: int):
    """Format column as percentage (0.0%)."""
    for row_idx in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row_idx, column=column)
        cell.number_format = '0.0%'


def add_total_row(worksheet, row: int, columns_to_sum: List[int], label_col: int = 1, label: str = "TOTAL"):
    """Add totals row with SUM formulas."""
    # Add label
    label_cell = worksheet.cell(row=row, column=label_col)
    label_cell.value = label
    label_cell.font = Font(name='Calibri', size=10, bold=True)
    label_cell.fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type='solid')

    # Add SUM formulas
    for col in columns_to_sum:
        cell = worksheet.cell(row=row, column=col)
        # Find first data row (assuming header is row 1)
        first_data_row = 2
        last_data_row = row - 1
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type='solid')


def auto_fit_columns(worksheet, max_width: int = 50):
    """Auto-fit column widths based on content."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_production_schedule(
    production_schedule,  # ProductionSchedule object
    labor_data: Optional[pd.DataFrame],
    output_path: str,
    cost_breakdown=None  # Optional TotalCostBreakdown
) -> str:
    """
    Export production schedule to formatted Excel file.

    Creates 4 sheets:
    1. Production Schedule - Detailed batch listing
    2. Daily Summary - Aggregated daily metrics
    3. Product Summary - Aggregated product metrics
    4. Metadata - Planning information

    Args:
        production_schedule: ProductionSchedule object with batches
        labor_data: DataFrame with labor hours by date (optional)
        output_path: Path to save Excel file
        cost_breakdown: Optional TotalCostBreakdown for cost information

    Returns:
        Path to created file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Production Schedule
    ws1 = wb.create_sheet("Production Schedule")

    # Prepare data
    schedule_data = []
    for batch in production_schedule.production_batches:
        schedule_data.append({
            'Date': batch.production_date,
            'Day of Week': batch.production_date.strftime('%A'),
            'Product Code': batch.product_id,
            'Product Name': batch.product_id,  # Could be enhanced with product names
            'Quantity (units)': batch.quantity,
            'Batch ID': batch.id,
            'Labor Hours Required': batch.labor_hours_used,
            'Start Time': '',  # Not yet implemented
            'End Time': '',  # Not yet implemented
            'Notes': f"Seq {batch.sequence_number}" if batch.sequence_number else ''
        })

    df_schedule = pd.DataFrame(schedule_data)

    # Only sort if we have data
    if len(df_schedule) > 0:
        df_schedule = df_schedule.sort_values(['Date', 'Product Code'])

    # Write headers
    headers = ['Date', 'Day of Week', 'Product Code', 'Product Name', 'Quantity (units)',
                'Batch ID', 'Labor Hours Required', 'Start Time', 'End Time', 'Notes']

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data
    for row_idx, row_data in enumerate(df_schedule.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Format date column
            if col_idx == 1 and isinstance(value, (datetime, pd.Timestamp)):
                cell.number_format = 'ddd, mmm dd'
            # Format quantity column
            elif col_idx == 5:
                cell.number_format = '#,##0'
            # Format labor hours column
            elif col_idx == 7:
                cell.number_format = '0.00'

    # Apply alternating rows
    if len(df_schedule) > 0:
        apply_alternating_rows(ws1, 2, len(df_schedule) + 1, 1, len(headers))

    # Add filters
    add_filters(ws1, len(headers))

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # Auto-fit columns
    auto_fit_columns(ws1)

    # Sheet 2: Daily Summary
    ws2 = wb.create_sheet("Daily Summary")

    # Aggregate by date
    daily_summary = []
    for prod_date, total_units in sorted(production_schedule.daily_totals.items()):
        labor_hours = production_schedule.daily_labor_hours.get(prod_date, 0.0)

        # Calculate capacity utilization
        max_daily_capacity = 19600  # 14 hours * 1400 units/hour
        capacity_util = (total_units / max_daily_capacity) if max_daily_capacity > 0 else 0

        # Determine overtime hours (assuming 12h fixed, >12h is OT)
        overtime_hours = max(0, labor_hours - 12.0)

        daily_summary.append({
            'Date': prod_date,
            'Day': prod_date.strftime('%A'),
            'Total Units': total_units,
            'Labor Hours': labor_hours,
            'Capacity Utilization': capacity_util,
            'Overtime Hours': overtime_hours,
            'Notes': 'Overtime' if overtime_hours > 0 else ''
        })

    df_daily = pd.DataFrame(daily_summary)

    # Write headers
    daily_headers = ['Date', 'Day', 'Total Units', 'Labor Hours', 'Capacity Utilization %', 'Overtime Hours', 'Notes']

    for col_idx, header in enumerate(daily_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data with conditional formatting
    for row_idx, row in enumerate(df_daily.itertuples(index=False), 2):
        capacity_util = row[4]  # Capacity Utilization
        overtime_hours = row[5]  # Overtime Hours

        # Determine row color
        row_color = None
        if capacity_util > 1.0:
            row_color = OVERLOAD_COLOR  # Red - over capacity
        elif overtime_hours > 0:
            row_color = WARNING_COLOR  # Orange - overtime
        elif capacity_util < 0.5:
            row_color = LOW_UTIL_COLOR  # Yellow - underutilization

        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Apply conditional formatting color
            if row_color:
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')

            # Format specific columns
            if col_idx == 1 and isinstance(value, (datetime, pd.Timestamp)):
                cell.number_format = 'ddd, mmm dd'
            elif col_idx == 3:  # Total Units
                cell.number_format = '#,##0'
            elif col_idx == 4:  # Labor Hours
                cell.number_format = '0.0'
            elif col_idx == 5:  # Capacity Utilization
                cell.number_format = '0.0%'
            elif col_idx == 6:  # Overtime Hours
                cell.number_format = '0.0'

    # Add totals row
    if len(df_daily) > 0:
        total_row = len(df_daily) + 2
        add_total_row(ws2, total_row, [3, 4, 6], label_col=2, label="TOTAL")

        # Format total row
        format_number(ws2, 3, total_row, total_row)
        ws2.cell(row=total_row, column=4).number_format = '0.0'
        ws2.cell(row=total_row, column=6).number_format = '0.0'

    # Auto-fit columns
    auto_fit_columns(ws2)

    # Sheet 3: Product Summary
    ws3 = wb.create_sheet("Product Summary")

    # Aggregate by product
    product_summary = {}
    for batch in production_schedule.production_batches:
        if batch.product_id not in product_summary:
            product_summary[batch.product_id] = {
                'quantity': 0,
                'production_days': set(),
                'labor_hours': 0
            }

        product_summary[batch.product_id]['quantity'] += batch.quantity
        product_summary[batch.product_id]['production_days'].add(batch.production_date)
        product_summary[batch.product_id]['labor_hours'] += batch.labor_hours_used

    product_data = []
    for product_id, stats in sorted(product_summary.items()):
        num_days = len(stats['production_days'])
        avg_batch = stats['quantity'] / num_days if num_days > 0 else 0

        product_data.append({
            'Product': product_id,
            'Total Quantity': stats['quantity'],
            '# Production Days': num_days,
            'Avg Batch Size': avg_batch,
            'Total Labor Hours': stats['labor_hours']
        })

    df_product = pd.DataFrame(product_data)

    # Write headers
    product_headers = ['Product', 'Total Quantity', '# Production Days', 'Avg Batch Size', 'Total Labor Hours']

    for col_idx, header in enumerate(product_headers, 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data
    for row_idx, row in enumerate(df_product.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Format numbers
            if col_idx in [2, 3]:  # Quantity columns
                cell.number_format = '#,##0'
            elif col_idx == 4:  # Avg batch
                cell.number_format = '#,##0.0'
            elif col_idx == 5:  # Labor hours
                cell.number_format = '0.0'

    # Apply alternating rows
    if len(df_product) > 0:
        apply_alternating_rows(ws3, 2, len(df_product) + 1, 1, len(product_headers))

    # Add bar chart
    if len(df_product) > 0:
        chart = BarChart()
        chart.title = "Production Quantity by Product"
        chart.x_axis.title = "Product"
        chart.y_axis.title = "Quantity (units)"

        data = Reference(ws3, min_col=2, min_row=1, max_row=len(df_product) + 1)
        categories = Reference(ws3, min_col=1, min_row=2, max_row=len(df_product) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 10
        chart.width = 20

        ws3.add_chart(chart, f"A{len(df_product) + 4}")

    # Auto-fit columns
    auto_fit_columns(ws3)

    # Sheet 4: Metadata
    ws4 = wb.create_sheet("Metadata")

    metadata = [
        ['Export Information', ''],
        ['Export Date & Time', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['Planning Period', ''],
        ['Start Date', production_schedule.schedule_start_date.strftime('%Y-%m-%d')],
        ['End Date', production_schedule.schedule_end_date.strftime('%Y-%m-%d')],
        ['', ''],
        ['Production Summary', ''],
        ['Total Production (units)', production_schedule.total_units],
        ['Total Labor Hours', production_schedule.total_labor_hours],
        ['Number of Batches', len(production_schedule.production_batches)],
        ['Number of Production Days', len(production_schedule.daily_totals)],
        ['', ''],
        ['Feasibility', ''],
        ['Status', 'FEASIBLE' if production_schedule.is_feasible() else 'INFEASIBLE'],
        ['Issues', len(production_schedule.infeasibilities)],
    ]

    # Add cost information if available
    if cost_breakdown:
        metadata.extend([
            ['', ''],
            ['Cost Summary', ''],
            ['Total Cost', f'${cost_breakdown.total_cost:,.2f}'],
            ['Labor Cost', f'${cost_breakdown.labor.total_cost:,.2f}'],
            ['Production Cost', f'${cost_breakdown.production.total_cost:,.2f}'],
            ['Transport Cost', f'${cost_breakdown.transport.total_cost:,.2f}'],
            ['Waste Cost', f'${cost_breakdown.waste.total_cost:,.2f}'],
            ['Cost Per Unit', f'${cost_breakdown.cost_per_unit_delivered:.4f}'],
        ])

    # Write metadata
    for row_idx, (label, value) in enumerate(metadata, 1):
        ws4.cell(row=row_idx, column=1).value = label
        ws4.cell(row=row_idx, column=2).value = value

        # Bold section headers
        if value == '' and label != '':
            ws4.cell(row=row_idx, column=1).font = Font(name='Calibri', size=12, bold=True)

    # Format numbers
    ws4.cell(row=9, column=2).number_format = '#,##0'
    ws4.cell(row=10, column=2).number_format = '0.0'

    # Auto-fit columns
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 30

    # Save workbook
    wb.save(output_path)
    return output_path


def export_cost_breakdown(
    cost_data,  # TotalCostBreakdown object
    output_path: str
) -> str:
    """
    Export cost breakdown to formatted Excel file.

    Creates 4 sheets:
    1. Cost Summary - High-level cost components
    2. Labor Cost Detail - Daily labor cost breakdown
    3. Transport Cost Detail - Cost by route
    4. Waste Cost Detail - Waste and shortage details

    Args:
        cost_data: TotalCostBreakdown object with all cost components
        output_path: Path to save Excel file

    Returns:
        Path to created file
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Cost Summary
    ws1 = wb.create_sheet("Cost Summary")

    summary_data = [
        ['Component', 'Amount ($)', '% of Total', 'Notes'],
        ['Labor', cost_data.labor.total_cost,
         cost_data.labor.total_cost / cost_data.total_cost if cost_data.total_cost > 0 else 0,
         f"{cost_data.labor.total_hours:.1f} hours"],
        ['Production', cost_data.production.total_cost,
         cost_data.production.total_cost / cost_data.total_cost if cost_data.total_cost > 0 else 0,
         f"{cost_data.production.total_units_produced:,.0f} units"],
        ['Transport', cost_data.transport.total_cost,
         cost_data.transport.total_cost / cost_data.total_cost if cost_data.total_cost > 0 else 0,
         f"{cost_data.transport.total_units_shipped:,.0f} units shipped"],
        ['Waste', cost_data.waste.total_cost,
         cost_data.waste.total_cost / cost_data.total_cost if cost_data.total_cost > 0 else 0,
         f"{cost_data.waste.expired_units + cost_data.waste.unmet_demand_units:,.0f} units"],
        ['TOTAL', cost_data.total_cost, 1.0,
         f"${cost_data.cost_per_unit_delivered:.4f} per unit"],
    ]

    # Write data
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Header row
            if row_idx == 1:
                style = create_header_style()
                cell.font = style['font']
                cell.fill = style['fill']
                cell.alignment = style['alignment']
                cell.border = style['border']
            # Total row
            elif row_idx == len(summary_data):
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type='solid')

            # Format columns
            if row_idx > 1 and col_idx == 2:
                cell.number_format = '$#,##0.00'
            elif row_idx > 1 and col_idx == 3:
                cell.number_format = '0.0%'

    # Highlight largest cost driver
    max_cost_row = 2
    max_cost = cost_data.labor.total_cost
    if cost_data.production.total_cost > max_cost:
        max_cost = cost_data.production.total_cost
        max_cost_row = 3
    if cost_data.transport.total_cost > max_cost:
        max_cost = cost_data.transport.total_cost
        max_cost_row = 4

    for col_idx in range(1, 5):
        ws1.cell(row=max_cost_row, column=col_idx).fill = PatternFill(
            start_color=HIGH_UTIL_COLOR, end_color=HIGH_UTIL_COLOR, fill_type='solid'
        )

    # Auto-fit columns
    auto_fit_columns(ws1)

    # Add pie chart
    chart = PieChart()
    chart.title = "Cost Breakdown"
    chart.height = 12
    chart.width = 16

    # Data for chart (exclude Total row)
    data = Reference(ws1, min_col=2, min_row=1, max_row=5)
    categories = Reference(ws1, min_col=1, min_row=2, max_row=5)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws1.add_chart(chart, "F2")

    # Sheet 2: Labor Cost Detail
    ws2 = wb.create_sheet("Labor Cost Detail")

    # Prepare daily labor data
    labor_rows = []
    for prod_date, daily_cost in sorted(cost_data.labor.daily_breakdown.items()):
        labor_rows.append({
            'Date': prod_date,
            'Day': prod_date.strftime('%A'),
            'Fixed Hours': daily_cost.get('fixed_hours', 0),
            'Overtime Hours': daily_cost.get('overtime_hours', 0),
            'Non-Fixed Hours': daily_cost.get('non_fixed_hours', 0),
            'Fixed Cost': daily_cost.get('fixed_cost', 0),
            'Overtime Cost': daily_cost.get('overtime_cost', 0),
            'Non-Fixed Cost': daily_cost.get('non_fixed_cost', 0),
            'Total': daily_cost.get('total_cost', 0),
        })

    df_labor = pd.DataFrame(labor_rows)

    # Write headers
    labor_headers = ['Date', 'Day', 'Fixed Hours', 'Overtime Hours', 'Non-Fixed Hours',
                     'Fixed Cost', 'Overtime Cost', 'Non-Fixed Cost', 'Total']

    for col_idx, header in enumerate(labor_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data
    if len(df_labor) > 0:
        for row_idx, row in enumerate(df_labor.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format columns
                if col_idx == 1:  # Date
                    cell.number_format = 'ddd, mmm dd'
                elif col_idx in [3, 4, 5]:  # Hours
                    cell.number_format = '0.0'
                elif col_idx in [6, 7, 8, 9]:  # Costs
                    cell.number_format = '$#,##0.00'

        # Apply alternating rows
        apply_alternating_rows(ws2, 2, len(df_labor) + 1, 1, len(labor_headers))

        # Add totals row
        total_row = len(df_labor) + 2
        add_total_row(ws2, total_row, [3, 4, 5, 6, 7, 8, 9], label_col=2, label="TOTAL")

        # Format totals
        for col in [3, 4, 5]:
            ws2.cell(row=total_row, column=col).number_format = '0.0'
        for col in [6, 7, 8, 9]:
            ws2.cell(row=total_row, column=col).number_format = '$#,##0.00'

        # Add line chart
        chart = LineChart()
        chart.title = "Labor Cost Over Time"
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Cost ($)"
        chart.height = 10
        chart.width = 20

        data = Reference(ws2, min_col=9, min_row=1, max_row=len(df_labor) + 1)
        categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_labor) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws2.add_chart(chart, f"A{len(df_labor) + 5}")

    # Auto-fit columns
    auto_fit_columns(ws2)

    # Sheet 3: Transport Cost Detail
    ws3 = wb.create_sheet("Transport Cost Detail")

    # Prepare transport data by route
    transport_rows = []
    for route, cost in sorted(cost_data.transport.cost_by_route.items(), key=lambda x: x[1], reverse=True):
        # Calculate units shipped on this route
        units_shipped = sum(
            detail['quantity'] for detail in cost_data.transport.shipment_details
            if detail.get('route') == route
        )

        cost_per_unit = cost / units_shipped if units_shipped > 0 else 0
        pct_of_transport = cost / cost_data.transport.total_cost if cost_data.transport.total_cost > 0 else 0

        # Parse route (format: "loc1 → loc2 → loc3")
        route_parts = route.split(' → ')
        origin = route_parts[0] if len(route_parts) > 0 else ''
        destination = route_parts[-1] if len(route_parts) > 0 else ''

        transport_rows.append({
            'Route': route,
            'Origin': origin,
            'Destination': destination,
            'Units Shipped': units_shipped,
            'Cost/Unit': cost_per_unit,
            'Total Cost': cost,
            '% of Transport Cost': pct_of_transport
        })

    df_transport = pd.DataFrame(transport_rows)

    # Write headers
    transport_headers = ['Route', 'Origin', 'Destination', 'Units Shipped', 'Cost/Unit', 'Total Cost', '% of Transport Cost']

    for col_idx, header in enumerate(transport_headers, 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data
    if len(df_transport) > 0:
        for row_idx, row in enumerate(df_transport.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws3.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format columns
                if col_idx == 4:  # Units
                    cell.number_format = '#,##0'
                elif col_idx in [5, 6]:  # Costs
                    cell.number_format = '$#,##0.00'
                elif col_idx == 7:  # Percentage
                    cell.number_format = '0.0%'

        # Apply alternating rows
        apply_alternating_rows(ws3, 2, len(df_transport) + 1, 1, len(transport_headers))

        # Add bar chart for top 10 routes
        chart = BarChart()
        chart.title = "Top 10 Routes by Cost"
        chart.x_axis.title = "Route"
        chart.y_axis.title = "Cost ($)"
        chart.height = 12
        chart.width = 20
        chart.type = "col"

        max_rows = min(11, len(df_transport) + 1)  # Top 10 + header
        data = Reference(ws3, min_col=6, min_row=1, max_row=max_rows)
        categories = Reference(ws3, min_col=1, min_row=2, max_row=max_rows)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws3.add_chart(chart, f"A{len(df_transport) + 4}")

    # Auto-fit columns
    auto_fit_columns(ws3)

    # Sheet 4: Waste Cost Detail
    ws4 = wb.create_sheet("Waste Cost Detail")

    if cost_data.waste.total_cost > 0:
        # Summary section
        ws4.cell(row=1, column=1).value = "Waste Summary"
        ws4.cell(row=1, column=1).font = Font(name='Calibri', size=12, bold=True)

        summary = [
            ['', ''],
            ['Total Waste Cost', f'${cost_data.waste.total_cost:,.2f}'],
            ['Expired Units', f'{cost_data.waste.expired_units:,.0f}'],
            ['Expired Cost', f'${cost_data.waste.expired_cost:,.2f}'],
            ['Unmet Demand Units', f'{cost_data.waste.unmet_demand_units:,.0f}'],
            ['Unmet Demand Cost', f'${cost_data.waste.unmet_demand_cost:,.2f}'],
        ]

        for row_idx, (label, value) in enumerate(summary, 2):
            ws4.cell(row=row_idx, column=1).value = label
            ws4.cell(row=row_idx, column=2).value = value

            if label:
                ws4.cell(row=row_idx, column=1).font = Font(name='Calibri', size=10, bold=True)

        # Waste by location
        if cost_data.waste.waste_by_location:
            ws4.cell(row=9, column=1).value = "Waste by Location"
            ws4.cell(row=9, column=1).font = Font(name='Calibri', size=12, bold=True)

            ws4.cell(row=10, column=1).value = "Location"
            ws4.cell(row=10, column=2).value = "Cost"

            style = create_header_style()
            for col in [1, 2]:
                cell = ws4.cell(row=10, column=col)
                cell.font = style['font']
                cell.fill = style['fill']
                cell.alignment = style['alignment']

            row_idx = 11
            for location, cost in sorted(cost_data.waste.waste_by_location.items(), key=lambda x: x[1], reverse=True):
                ws4.cell(row=row_idx, column=1).value = location
                ws4.cell(row=row_idx, column=2).value = cost
                ws4.cell(row=row_idx, column=2).number_format = '$#,##0.00'
                row_idx += 1
    else:
        ws4.cell(row=1, column=1).value = "No Waste Costs"
        ws4.cell(row=1, column=1).font = Font(name='Calibri', size=12, bold=True, color='008000')
        ws4.cell(row=2, column=1).value = "All demand met with acceptable shelf life!"

    # Auto-fit columns
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 20

    # Save workbook
    wb.save(output_path)
    return output_path


def export_shipment_plan(
    shipment_data: List,  # List of Shipment objects
    truck_plan,  # TruckLoadPlan object
    output_path: str,
    truck_capacity: int = 14080  # 44 pallets * 320 units
) -> str:
    """
    Export shipment plan to formatted Excel file.

    Creates 4 sheets:
    1. Truck Loading Schedule - Detailed truck-by-truck loading
    2. Daily Shipments - Daily summary of shipments
    3. Destination Summary - Aggregated by destination
    4. Truck Manifests - One row per truck for printing

    Args:
        shipment_data: List of Shipment objects
        truck_plan: TruckLoadPlan object with truck assignments
        output_path: Path to save Excel file
        truck_capacity: Maximum units per truck (default 14,080)

    Returns:
        Path to created file
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Truck Loading Schedule
    ws1 = wb.create_sheet("Truck Loading Schedule")

    # Prepare truck loading data
    loading_data = []
    for load in sorted(truck_plan.loads, key=lambda x: (x.departure_date, x.departure_type)):
        for shipment in load.shipments:
            pallets = (shipment.quantity / 320)  # 320 units per pallet

            loading_data.append({
                'Departure Date': load.departure_date,
                'Truck Name': load.truck_name,
                'Origin': shipment.origin_id if hasattr(shipment, 'origin_id') else '',
                'Destination': load.destination_id or '',
                'Product': shipment.product_id,
                'Quantity (units)': shipment.quantity,
                'Pallets': pallets,
                'Truck % Full': load.capacity_utilization,
                'Arrival Date': shipment.delivery_date if hasattr(shipment, 'delivery_date') else ''
            })

    df_loading = pd.DataFrame(loading_data)

    # Write headers
    loading_headers = ['Departure Date', 'Truck Name', 'Origin', 'Destination', 'Product',
                       'Quantity (units)', 'Pallets', 'Truck % Full', 'Arrival Date']

    for col_idx, header in enumerate(loading_headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data with conditional formatting
    if len(df_loading) > 0:
        for row_idx, row in enumerate(df_loading.itertuples(index=False), 2):
            utilization = row[7]  # Truck % Full

            # Determine row color based on utilization
            row_color = None
            if utilization > 1.0:
                row_color = OVERLOAD_COLOR  # Red - overloaded
            elif utilization >= 0.8:
                row_color = HIGH_UTIL_COLOR  # Green - optimal
            elif utilization >= 0.5:
                row_color = None  # No color - acceptable
            else:
                row_color = LOW_UTIL_COLOR  # Yellow - underutilized

            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Apply conditional formatting
                if row_color:
                    cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')

                # Format columns
                if col_idx in [1, 9]:  # Dates
                    cell.number_format = 'yyyy-mm-dd'
                elif col_idx == 6:  # Quantity
                    cell.number_format = '#,##0'
                elif col_idx == 7:  # Pallets
                    cell.number_format = '0.0'
                elif col_idx == 8:  # % Full
                    cell.number_format = '0.0%'

        # Apply alternating rows
        apply_alternating_rows(ws1, 2, len(df_loading) + 1, 1, len(loading_headers))

    # Add filters
    add_filters(ws1, len(loading_headers))

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # Auto-fit columns
    auto_fit_columns(ws1)

    # Sheet 2: Daily Shipments
    ws2 = wb.create_sheet("Daily Shipments")

    # Aggregate by date
    daily_shipments = {}
    for load in truck_plan.loads:
        date_key = load.departure_date
        if date_key not in daily_shipments:
            daily_shipments[date_key] = {
                'trucks': 0,
                'units': 0,
                'pallets': 0,
                'destinations': set()
            }

        daily_shipments[date_key]['trucks'] += 1
        daily_shipments[date_key]['units'] += load.total_units
        daily_shipments[date_key]['pallets'] += load.total_pallets
        if load.destination_id:
            daily_shipments[date_key]['destinations'].add(load.destination_id)

    daily_data = []
    for date_key, stats in sorted(daily_shipments.items()):
        daily_data.append({
            'Date': date_key,
            '# Trucks': stats['trucks'],
            'Total Units': stats['units'],
            'Total Pallets': stats['pallets'],
            'Destinations': ', '.join(sorted(stats['destinations'])),
            'Notes': ''
        })

    df_daily = pd.DataFrame(daily_data)

    # Write headers
    daily_headers = ['Date', '# Trucks', 'Total Units', 'Total Pallets', 'Destinations', 'Notes']

    for col_idx, header in enumerate(daily_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data
    if len(df_daily) > 0:
        for row_idx, row in enumerate(df_daily.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format columns
                if col_idx == 1:  # Date
                    cell.number_format = 'ddd, mmm dd'
                elif col_idx in [2, 3, 4]:  # Numbers
                    cell.number_format = '#,##0'

        # Apply alternating rows
        apply_alternating_rows(ws2, 2, len(df_daily) + 1, 1, len(daily_headers))

    # Auto-fit columns
    auto_fit_columns(ws2)

    # Sheet 3: Destination Summary
    ws3 = wb.create_sheet("Destination Summary")

    # Aggregate by destination
    dest_summary = {}
    for shipment in shipment_data:
        dest_id = shipment.destination_id
        if dest_id not in dest_summary:
            dest_summary[dest_id] = {
                'units': 0,
                'deliveries': 0,
                'dates': []
            }

        dest_summary[dest_id]['units'] += shipment.quantity
        dest_summary[dest_id]['deliveries'] += 1
        if hasattr(shipment, 'delivery_date'):
            dest_summary[dest_id]['dates'].append(shipment.delivery_date)

    dest_data = []
    for dest_id, stats in sorted(dest_summary.items()):
        avg_delivery = stats['units'] / stats['deliveries'] if stats['deliveries'] > 0 else 0
        first_delivery = min(stats['dates']) if stats['dates'] else None
        last_delivery = max(stats['dates']) if stats['dates'] else None

        dest_data.append({
            'Destination': dest_id,
            'Total Units Received': stats['units'],
            '# Deliveries': stats['deliveries'],
            'Avg Delivery Size': avg_delivery,
            'First Delivery': first_delivery,
            'Last Delivery': last_delivery
        })

    df_dest = pd.DataFrame(dest_data)

    # Write headers
    dest_headers = ['Destination', 'Total Units Received', '# Deliveries', 'Avg Delivery Size',
                    'First Delivery', 'Last Delivery']

    for col_idx, header in enumerate(dest_headers, 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data
    if len(df_dest) > 0:
        for row_idx, row in enumerate(df_dest.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws3.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format columns
                if col_idx in [2, 3, 4]:  # Numbers
                    cell.number_format = '#,##0'
                elif col_idx in [5, 6]:  # Dates
                    if value:
                        cell.number_format = 'yyyy-mm-dd'

        # Apply alternating rows
        apply_alternating_rows(ws3, 2, len(df_dest) + 1, 1, len(dest_headers))

    # Auto-fit columns
    auto_fit_columns(ws3)

    # Sheet 4: Truck Manifests
    ws4 = wb.create_sheet("Truck Manifests")

    # One row per truck
    manifest_data = []
    for load in sorted(truck_plan.loads, key=lambda x: (x.departure_date, x.departure_type)):
        # Aggregate products on this truck
        products = {}
        for shipment in load.shipments:
            if shipment.product_id not in products:
                products[shipment.product_id] = 0
            products[shipment.product_id] += shipment.quantity

        products_list = ', '.join([f"{pid} ({qty:.0f})" for pid, qty in sorted(products.items())])

        manifest_data.append({
            'Truck ID': load.truck_schedule_id,
            'Date': load.departure_date,
            'Route': f"{load.departure_type.title()} to {load.destination_id or 'Multiple'}",
            'Products': products_list,
            'Total Units': load.total_units,
            'Total Pallets': load.total_pallets,
            'Utilization %': load.capacity_utilization
        })

    df_manifest = pd.DataFrame(manifest_data)

    # Write headers
    manifest_headers = ['Truck ID', 'Date', 'Route', 'Products', 'Total Units', 'Total Pallets', 'Utilization %']

    for col_idx, header in enumerate(manifest_headers, 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.value = header
        style = create_header_style()
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # Write data
    if len(df_manifest) > 0:
        for row_idx, row in enumerate(df_manifest.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws4.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format columns
                if col_idx == 2:  # Date
                    cell.number_format = 'yyyy-mm-dd'
                elif col_idx in [5, 6]:  # Numbers
                    cell.number_format = '#,##0'
                elif col_idx == 7:  # Percentage
                    cell.number_format = '0.0%'

        # Apply alternating rows
        apply_alternating_rows(ws4, 2, len(df_manifest) + 1, 1, len(manifest_headers))

    # Auto-fit columns
    auto_fit_columns(ws4)

    # Save workbook
    wb.save(output_path)
    return output_path
