"""
Tests for Excel export templates.

Tests the three main export functions:
- export_production_schedule()
- export_cost_breakdown()
- export_shipment_plan()
"""

import pytest
import tempfile
import os
from datetime import date, time, timedelta
from pathlib import Path
from openpyxl import load_workbook

from src.exporters import (
    export_production_schedule,
    export_cost_breakdown,
    export_shipment_plan,
)
from src.production.scheduler import ProductionSchedule
from src.models.production_batch import ProductionBatch
from src.models.product import ProductState
from src.models.shipment import Shipment
from src.distribution.truck_loader import TruckLoad, TruckLoadPlan
from src.costs.cost_breakdown import (
    TotalCostBreakdown,
    LaborCostBreakdown,
    ProductionCostBreakdown,
    TransportCostBreakdown,
    WasteCostBreakdown,
)


@pytest.fixture
def temp_output_path():
    """Create temporary file path for test outputs."""
    temp_dir = tempfile.gettempdir()
    filename = f"test_export_{os.getpid()}.xlsx"
    path = os.path.join(temp_dir, filename)
    yield path
    # Cleanup
    if os.path.exists(path):
        os.remove(path)


@pytest.fixture
def sample_production_schedule():
    """Create sample production schedule for testing."""
    batches = [
        ProductionBatch(
            id="BATCH-001",
            product_id="PROD_A",
            manufacturing_site_id="6122",
            production_date=date(2025, 10, 15),
            quantity=1000,
            initial_state=ProductState.AMBIENT,
            labor_hours_used=0.71,
            production_cost=500.0,
            sequence_number=1,
            changeover_time_hours=0.0,
        ),
        ProductionBatch(
            id="BATCH-002",
            product_id="PROD_B",
            manufacturing_site_id="6122",
            production_date=date(2025, 10, 15),
            quantity=1500,
            initial_state=ProductState.AMBIENT,
            labor_hours_used=1.07,
            production_cost=750.0,
            sequence_number=2,
            changeover_time_hours=0.5,
        ),
        ProductionBatch(
            id="BATCH-003",
            product_id="PROD_A",
            manufacturing_site_id="6122",
            production_date=date(2025, 10, 16),
            quantity=2000,
            initial_state=ProductState.AMBIENT,
            labor_hours_used=1.43,
            production_cost=1000.0,
            sequence_number=1,
            changeover_time_hours=0.0,
        ),
    ]

    daily_totals = {
        date(2025, 10, 15): 2500,
        date(2025, 10, 16): 2000,
    }

    daily_labor_hours = {
        date(2025, 10, 15): 1.78,
        date(2025, 10, 16): 1.43,
    }

    return ProductionSchedule(
        manufacturing_site_id="6122",
        schedule_start_date=date(2025, 10, 15),
        schedule_end_date=date(2025, 10, 16),
        production_batches=batches,
        daily_totals=daily_totals,
        daily_labor_hours=daily_labor_hours,
        infeasibilities=[],
        total_units=4500,
        total_labor_hours=3.21,
        requirements=[],
    )


@pytest.fixture
def sample_cost_breakdown():
    """Create sample cost breakdown for testing."""
    # Labor costs
    labor = LaborCostBreakdown(
        total_cost=1500.0,
        fixed_hours_cost=1000.0,
        overtime_cost=300.0,
        non_fixed_labor_cost=200.0,
        total_hours=50.0,
        fixed_hours=40.0,
        overtime_hours=8.0,
        non_fixed_hours=2.0,
        daily_breakdown={
            date(2025, 10, 15): {
                'fixed_hours': 12.0,
                'overtime_hours': 2.0,
                'non_fixed_hours': 0.0,
                'fixed_cost': 400.0,
                'overtime_cost': 100.0,
                'non_fixed_cost': 0.0,
                'total_cost': 500.0,
            },
            date(2025, 10, 16): {
                'fixed_hours': 12.0,
                'overtime_hours': 0.0,
                'non_fixed_hours': 0.0,
                'fixed_cost': 400.0,
                'overtime_cost': 0.0,
                'non_fixed_cost': 0.0,
                'total_cost': 400.0,
            },
        },
    )

    # Production costs
    production = ProductionCostBreakdown(
        total_cost=2250.0,
        total_units_produced=4500,
        average_cost_per_unit=0.5,
        cost_by_product={
            'PROD_A': 1500.0,
            'PROD_B': 750.0,
        },
        cost_by_date={
            date(2025, 10, 15): 1250.0,
            date(2025, 10, 16): 1000.0,
        },
        batch_details=[
            {'batch_id': 'BATCH-001', 'product_id': 'PROD_A', 'quantity': 1000, 'cost': 500.0},
            {'batch_id': 'BATCH-002', 'product_id': 'PROD_B', 'quantity': 1500, 'cost': 750.0},
            {'batch_id': 'BATCH-003', 'product_id': 'PROD_A', 'quantity': 2000, 'cost': 1000.0},
        ],
    )

    # Transport costs
    transport = TransportCostBreakdown(
        total_cost=900.0,
        total_units_shipped=4500,
        average_cost_per_unit=0.2,
        cost_by_route={
            '6122 → 6125': 500.0,
            '6122 → 6104': 400.0,
        },
        cost_by_leg={
            '6122 → 6125': 500.0,
            '6122 → 6104': 400.0,
        },
        shipment_details=[
            {'route': '6122 → 6125', 'quantity': 2500, 'cost': 500.0},
            {'route': '6122 → 6104', 'quantity': 2000, 'cost': 400.0},
        ],
    )

    # Waste costs
    waste = WasteCostBreakdown(
        total_cost=50.0,
        expired_units=100,
        expired_cost=50.0,
        unmet_demand_units=0,
        unmet_demand_cost=0.0,
        waste_by_location={
            '6125': 30.0,
            '6104': 20.0,
        },
        waste_by_product={},
        waste_details=[],
    )

    # Total breakdown
    total = TotalCostBreakdown(
        total_cost=4700.0,
        labor=labor,
        production=production,
        transport=transport,
        waste=waste,
        cost_per_unit_delivered=1.044,
    )

    return total


@pytest.fixture
def sample_shipments_and_trucks():
    """Create sample shipments and truck plan for testing."""
    from src.models.route import Route
    from src.models.location import StorageMode

    # Create sample route
    route_6125 = Route(
        id="R1",
        origin_id="6122",
        destination_id="6125",
        transport_mode=StorageMode.AMBIENT,
        transit_time_days=1,
        cost=0.2,
    )

    # Create shipments
    shipments = [
        Shipment(
            id="SHIP-001",
            batch_id="BATCH-001",
            product_id="PROD_A",
            origin_id="6122",
            destination_id="6125",
            quantity=1000,
            delivery_date=date(2025, 10, 16),
            production_date=date(2025, 10, 15),
            route=route_6125,
        ),
        Shipment(
            id="SHIP-002",
            batch_id="BATCH-002",
            product_id="PROD_B",
            origin_id="6122",
            destination_id="6125",
            quantity=1500,
            delivery_date=date(2025, 10, 16),
            production_date=date(2025, 10, 15),
            route=route_6125,
        ),
    ]

    # Create truck loads
    load1 = TruckLoad(
        truck_schedule_id="TRUCK-001",
        truck_name="Morning Truck 6125",
        departure_date=date(2025, 10, 15),
        departure_type="morning",
        departure_time=time(6, 0),
        destination_id="6125",
        shipments=shipments,
        total_units=2500,
        total_pallets=8,
        capacity_units=14080,
        capacity_pallets=44,
        capacity_utilization=0.18,
        utilization_pct=0.18,
    )

    truck_plan = TruckLoadPlan(
        loads=[load1],
        unassigned_shipments=[],
        infeasibilities=[],
        total_trucks_used=1,
        total_shipments=2,
        average_utilization=0.18,
    )

    return shipments, truck_plan


class TestProductionScheduleExport:
    """Test production schedule Excel export."""

    def test_export_basic_schedule(self, sample_production_schedule, temp_output_path):
        """Test basic production schedule export."""
        result_path = export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        assert result_path == temp_output_path
        assert os.path.exists(result_path)

        # Load workbook and verify sheets
        wb = load_workbook(result_path)
        assert "Production Schedule" in wb.sheetnames
        assert "Daily Summary" in wb.sheetnames
        assert "Product Summary" in wb.sheetnames
        assert "Metadata" in wb.sheetnames

    def test_production_schedule_sheet_content(self, sample_production_schedule, temp_output_path):
        """Test Production Schedule sheet has correct content."""
        export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Production Schedule"]

        # Check headers
        assert ws['A1'].value == 'Date'
        assert ws['C1'].value == 'Product Code'
        assert ws['E1'].value == 'Quantity (units)'

        # Check data rows (3 batches + 1 header)
        assert ws.max_row == 4

        # Check first data row
        assert ws['C2'].value == 'PROD_A'
        assert ws['E2'].value == 1000

    def test_daily_summary_sheet(self, sample_production_schedule, temp_output_path):
        """Test Daily Summary sheet has correct aggregations."""
        export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Daily Summary"]

        # Check headers
        assert ws['A1'].value == 'Date'
        assert ws['C1'].value == 'Total Units'
        assert ws['D1'].value == 'Labor Hours'

        # Should have 2 days of data + header
        assert ws.max_row >= 3

    def test_product_summary_sheet(self, sample_production_schedule, temp_output_path):
        """Test Product Summary sheet has correct aggregations."""
        export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Product Summary"]

        # Check headers
        assert ws['A1'].value == 'Product'
        assert ws['B1'].value == 'Total Quantity'

        # Should have 2 products + header
        assert ws.max_row >= 3

    def test_metadata_sheet(self, sample_production_schedule, temp_output_path):
        """Test Metadata sheet has export information."""
        export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Metadata"]

        # Check key metadata fields
        assert 'Export Date & Time' in [cell.value for cell in ws['A']]
        assert 'Total Production (units)' in [cell.value for cell in ws['A']]
        assert 'Status' in [cell.value for cell in ws['A']]

    def test_export_with_cost_breakdown(self, sample_production_schedule, sample_cost_breakdown, temp_output_path):
        """Test export with cost breakdown included."""
        result_path = export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
            cost_breakdown=sample_cost_breakdown,
        )

        wb = load_workbook(result_path)
        ws = wb["Metadata"]

        # Check that cost information is present
        cells = [cell.value for cell in ws['A']]
        assert 'Total Cost' in cells


class TestCostBreakdownExport:
    """Test cost breakdown Excel export."""

    def test_export_basic_costs(self, sample_cost_breakdown, temp_output_path):
        """Test basic cost breakdown export."""
        result_path = export_cost_breakdown(
            cost_data=sample_cost_breakdown,
            output_path=temp_output_path,
        )

        assert result_path == temp_output_path
        assert os.path.exists(result_path)

        # Load workbook and verify sheets
        wb = load_workbook(result_path)
        assert "Cost Summary" in wb.sheetnames
        assert "Labor Cost Detail" in wb.sheetnames
        assert "Transport Cost Detail" in wb.sheetnames
        assert "Waste Cost Detail" in wb.sheetnames

    def test_cost_summary_sheet(self, sample_cost_breakdown, temp_output_path):
        """Test Cost Summary sheet content."""
        export_cost_breakdown(
            cost_data=sample_cost_breakdown,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Cost Summary"]

        # Check headers
        assert ws['A1'].value == 'Component'
        assert ws['B1'].value == 'Amount ($)'

        # Check components are listed
        components = [ws[f'A{i}'].value for i in range(2, 7)]
        assert 'Labor' in components
        assert 'Production' in components
        assert 'Transport' in components
        assert 'Waste' in components
        assert 'TOTAL' in components

    def test_labor_cost_detail_sheet(self, sample_cost_breakdown, temp_output_path):
        """Test Labor Cost Detail sheet."""
        export_cost_breakdown(
            cost_data=sample_cost_breakdown,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Labor Cost Detail"]

        # Check headers
        assert ws['A1'].value == 'Date'
        assert ws['C1'].value == 'Fixed Hours'
        assert ws['F1'].value == 'Fixed Cost'

        # Should have daily breakdowns
        assert ws.max_row >= 3

    def test_transport_cost_detail_sheet(self, sample_cost_breakdown, temp_output_path):
        """Test Transport Cost Detail sheet."""
        export_cost_breakdown(
            cost_data=sample_cost_breakdown,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Transport Cost Detail"]

        # Check headers
        assert ws['A1'].value == 'Route'
        assert ws['D1'].value == 'Units Shipped'
        assert ws['F1'].value == 'Total Cost'

    def test_number_formatting(self, sample_cost_breakdown, temp_output_path):
        """Test that currency and percentage formatting is applied."""
        export_cost_breakdown(
            cost_data=sample_cost_breakdown,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Cost Summary"]

        # Check currency format for cost column
        cost_cell = ws['B2']  # First cost value
        assert '$' in cost_cell.number_format or 'Currency' in str(cost_cell.number_format)


class TestShipmentPlanExport:
    """Test shipment plan Excel export."""

    def test_export_basic_shipments(self, sample_shipments_and_trucks, temp_output_path):
        """Test basic shipment plan export."""
        shipments, truck_plan = sample_shipments_and_trucks

        result_path = export_shipment_plan(
            shipment_data=shipments,
            truck_plan=truck_plan,
            output_path=temp_output_path,
        )

        assert result_path == temp_output_path
        assert os.path.exists(result_path)

        # Load workbook and verify sheets
        wb = load_workbook(result_path)
        assert "Truck Loading Schedule" in wb.sheetnames
        assert "Daily Shipments" in wb.sheetnames
        assert "Destination Summary" in wb.sheetnames
        assert "Truck Manifests" in wb.sheetnames

    def test_truck_loading_schedule_sheet(self, sample_shipments_and_trucks, temp_output_path):
        """Test Truck Loading Schedule sheet content."""
        shipments, truck_plan = sample_shipments_and_trucks

        export_shipment_plan(
            shipment_data=shipments,
            truck_plan=truck_plan,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Truck Loading Schedule"]

        # Check headers
        assert ws['A1'].value == 'Departure Date'
        assert ws['B1'].value == 'Truck Name'
        assert ws['D1'].value == 'Destination'
        assert ws['F1'].value == 'Quantity (units)'

    def test_daily_shipments_sheet(self, sample_shipments_and_trucks, temp_output_path):
        """Test Daily Shipments sheet aggregates correctly."""
        shipments, truck_plan = sample_shipments_and_trucks

        export_shipment_plan(
            shipment_data=shipments,
            truck_plan=truck_plan,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Daily Shipments"]

        # Check headers
        assert ws['A1'].value == 'Date'
        assert ws['B1'].value == '# Trucks'
        assert ws['C1'].value == 'Total Units'

    def test_destination_summary_sheet(self, sample_shipments_and_trucks, temp_output_path):
        """Test Destination Summary sheet."""
        shipments, truck_plan = sample_shipments_and_trucks

        export_shipment_plan(
            shipment_data=shipments,
            truck_plan=truck_plan,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Destination Summary"]

        # Check headers
        assert ws['A1'].value == 'Destination'
        assert ws['B1'].value == 'Total Units Received'

    def test_truck_manifests_sheet(self, sample_shipments_and_trucks, temp_output_path):
        """Test Truck Manifests sheet."""
        shipments, truck_plan = sample_shipments_and_trucks

        export_shipment_plan(
            shipment_data=shipments,
            truck_plan=truck_plan,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Truck Manifests"]

        # Check headers
        assert ws['A1'].value == 'Truck ID'
        assert ws['E1'].value == 'Total Units'
        assert ws['G1'].value == 'Utilization %'


class TestExportErrorHandling:
    """Test error handling in export functions."""

    def test_invalid_path_raises_error(self, sample_production_schedule):
        """Test that invalid output path raises appropriate error."""
        invalid_path = "/nonexistent/directory/file.xlsx"

        with pytest.raises(Exception):
            export_production_schedule(
                production_schedule=sample_production_schedule,
                labor_data=None,
                output_path=invalid_path,
            )

    def test_empty_schedule_exports(self, temp_output_path):
        """Test that empty schedule can be exported without errors."""
        empty_schedule = ProductionSchedule(
            manufacturing_site_id="6122",
            schedule_start_date=date.today(),
            schedule_end_date=date.today(),
            production_batches=[],
            daily_totals={},
            daily_labor_hours={},
            infeasibilities=[],
            total_units=0,
            total_labor_hours=0,
            requirements=[],
        )

        result_path = export_production_schedule(
            production_schedule=empty_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        assert os.path.exists(result_path)

    def test_empty_cost_breakdown_exports(self, temp_output_path):
        """Test that empty cost breakdown can be exported."""
        empty_costs = TotalCostBreakdown()

        result_path = export_cost_breakdown(
            cost_data=empty_costs,
            output_path=temp_output_path,
        )

        assert os.path.exists(result_path)


class TestExportFormatting:
    """Test Excel formatting features."""

    def test_alternating_row_colors(self, sample_production_schedule, temp_output_path):
        """Test that alternating row colors are applied."""
        export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Production Schedule"]

        # Check that some cells have fill color (alternating rows)
        row_2_fill = ws['A2'].fill
        row_3_fill = ws['A3'].fill

        # At least one should have a fill color
        assert (row_2_fill.start_color.rgb is not None or
                row_3_fill.start_color.rgb is not None)

    def test_header_formatting(self, sample_production_schedule, temp_output_path):
        """Test that headers have proper formatting."""
        export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Production Schedule"]

        # Check header cell formatting
        header_cell = ws['A1']
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb is not None  # Has background color

    def test_freeze_panes(self, sample_production_schedule, temp_output_path):
        """Test that freeze panes is set on header row."""
        export_production_schedule(
            production_schedule=sample_production_schedule,
            labor_data=None,
            output_path=temp_output_path,
        )

        wb = load_workbook(temp_output_path)
        ws = wb["Production Schedule"]

        # Check that freeze panes is set
        assert ws.freeze_panes is not None
        assert ws.freeze_panes == 'A2'  # Freeze at row 2


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
