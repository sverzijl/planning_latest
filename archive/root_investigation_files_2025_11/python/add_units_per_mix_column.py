#!/usr/bin/env python3
"""
Add units_per_mix column to all Network Config Excel files for mix-based production feature.

This script:
1. Processes all three Network_Config files:
   - Network_Config.xlsx (primary config)
   - Network_Config_Unified.xlsx (unified model config)
   - Network_Config_backup.xlsx (backup config)
2. Adds a "Products" sheet to each file if it doesn't exist
3. Populates products from forecast files
4. Adds units_per_mix column with realistic values
5. Saves the updated files
"""

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import pandas as pd

# File paths
NETWORK_CONFIG_FILES = [
    Path("data/examples/Network_Config.xlsx"),
    Path("data/examples/Network_Config_Unified.xlsx"),
    Path("data/examples/Network_Config_backup.xlsx"),
]
FORECAST_PATH = Path("data/examples/Gfree Forecast.xlsm")
FORECAST_LATEST_PATH = Path("data/examples/Gluten Free Forecast - Latest.xlsm")

# Realistic units_per_mix values for different products
# These should reflect actual batch sizes from manufacturing
# NOTE: We define all 5 products even though only 2 (G610, G142) are found in current forecast files.
# This ensures the Network Config files are complete and ready for any product mix scenarios
# that may be used in testing or production planning.
PRODUCT_MIX_SIZES = {
    "G142": 415,
    "G144": 387,
    "G147": 520,
    "G153": 450,
    "G610": 395,
}

# Default shelf life values (days)
DEFAULT_SHELF_LIFE = {
    "ambient": 17,
    "frozen": 120,
    "thawed": 14,
    "min_acceptable": 7,
}


def get_unique_products_from_forecast(forecast_path: Path) -> set:
    """Extract unique product IDs from forecast file."""
    try:
        # Try standard Forecast sheet first
        df = pd.read_excel(forecast_path, sheet_name="Forecast", engine="openpyxl")
        if "product_id" in df.columns:
            products = set(df["product_id"].unique())
            print(f"  Found {len(products)} products in Forecast sheet: {products}")
            return products
    except Exception as e:
        print(f"  Could not read standard Forecast sheet: {e}")

    # Try SAP IBP format (product codes in sheet names)
    try:
        wb = load_workbook(forecast_path, data_only=True, read_only=True)
        products = set()
        for sheet_name in wb.sheetnames:
            # SAP IBP sheets often have format like "G610_RET", "G142_RET", etc.
            if "_" in sheet_name:
                product_code = sheet_name.split("_")[0]
                if product_code.startswith("G"):
                    products.add(product_code)

        if products:
            print(f"  Found {len(products)} products from sheet names: {products}")
            return products
    except Exception as e:
        print(f"  Could not extract from sheet names: {e}")

    return set()


def create_products_sheet(wb: Workbook, products: set) -> None:
    """Create or update Products sheet in workbook."""

    # Create Products sheet if it doesn't exist
    if "Products" in wb.sheetnames:
        print("  Products sheet exists, will update it")
        ws = wb["Products"]
        ws.delete_rows(1, ws.max_row)
    else:
        print("  Creating new Products sheet")
        ws = wb.create_sheet("Products", 0)  # Insert at beginning

    # Header row styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Write headers
    headers = [
        "product_id",
        "name",
        "sku",
        "shelf_life_ambient_days",
        "shelf_life_frozen_days",
        "shelf_life_after_thaw_days",
        "min_acceptable_shelf_life_days",
        "units_per_mix"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write product data
    sorted_products = sorted(products)
    for row_idx, product_id in enumerate(sorted_products, start=2):
        # Get mix size from our mapping, or use a default
        units_per_mix = PRODUCT_MIX_SIZES.get(product_id, 400)

        # Create product row
        ws.cell(row=row_idx, column=1, value=product_id)  # product_id
        ws.cell(row=row_idx, column=2, value=f"Product {product_id}")  # name
        ws.cell(row=row_idx, column=3, value=product_id)  # sku
        ws.cell(row=row_idx, column=4, value=DEFAULT_SHELF_LIFE["ambient"])  # ambient shelf life
        ws.cell(row=row_idx, column=5, value=DEFAULT_SHELF_LIFE["frozen"])  # frozen shelf life
        ws.cell(row=row_idx, column=6, value=DEFAULT_SHELF_LIFE["thawed"])  # thawed shelf life
        ws.cell(row=row_idx, column=7, value=DEFAULT_SHELF_LIFE["min_acceptable"])  # min acceptable
        ws.cell(row=row_idx, column=8, value=units_per_mix)  # units_per_mix

        print(f"  Added product {product_id} with units_per_mix={units_per_mix}")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width

    print(f"  Products sheet created with {len(sorted_products)} products")


def main():
    """Main execution function."""
    print("=" * 80)
    print("Adding units_per_mix column to Excel files")
    print("=" * 80)

    # Step 1: Get unique products from forecast files
    print("\nStep 1: Extracting products from forecast files...")
    all_products = set()

    for forecast_path in [FORECAST_PATH, FORECAST_LATEST_PATH]:
        if forecast_path.exists():
            print(f"\nReading: {forecast_path}")
            products = get_unique_products_from_forecast(forecast_path)
            all_products.update(products)
        else:
            print(f"\nSkipping (not found): {forecast_path}")

    if not all_products:
        print("\n⚠️  No products found in forecast files!")
        print("Using default product set: G142, G144, G147, G153, G610")
        all_products = set(PRODUCT_MIX_SIZES.keys())

    print(f"\nTotal unique products found: {len(all_products)}")
    print(f"Products: {sorted(all_products)}")

    # Step 2: Update all Network_Config files
    print("\nStep 2: Updating Network_Config files...")

    updated_files = []
    skipped_files = []

    for config_path in NETWORK_CONFIG_FILES:
        print(f"\n  Processing: {config_path}")

        if not config_path.exists():
            print(f"    ⚠️  Skipping (not found)")
            skipped_files.append(config_path)
            continue

        try:
            # Load workbook
            wb = load_workbook(config_path)
            print(f"    Loaded workbook with sheets: {wb.sheetnames}")

            # Create/update Products sheet
            create_products_sheet(wb, all_products)

            # Save workbook
            wb.save(config_path)
            print(f"    ✅ Successfully updated {config_path.name}")
            updated_files.append(config_path)

        except Exception as e:
            print(f"    ❌ Error updating {config_path.name}: {e}")
            skipped_files.append(config_path)

    # Step 3: Verify the changes
    print("\nStep 3: Verifying changes...")

    for config_path in updated_files:
        print(f"\n  Verifying: {config_path.name}")
        wb_verify = load_workbook(config_path)

        if "Products" in wb_verify.sheetnames:
            ws = wb_verify["Products"]
            print(f"    ✓ Products sheet exists")
            print(f"    ✓ Contains {ws.max_row - 1} product rows")

            # Check headers
            headers = [cell.value for cell in ws[1]]
            if "units_per_mix" in headers:
                print(f"    ✓ units_per_mix column present at position {headers.index('units_per_mix') + 1}")
            else:
                print(f"    ❌ units_per_mix column NOT found!")

            # Display sample data
            print(f"    Sample product data:")
            for row_idx in range(2, min(5, ws.max_row + 1)):
                product_id = ws.cell(row=row_idx, column=1).value
                units_per_mix = ws.cell(row=row_idx, column=8).value
                print(f"      {product_id}: {units_per_mix} units/mix")

    # Summary
    print("\n" + "=" * 80)
    print("✅ COMPLETE - units_per_mix column added successfully!")
    print("=" * 80)
    print(f"\nFiles updated: {len(updated_files)}")
    for path in updated_files:
        print(f"  ✓ {path.name}")

    if skipped_files:
        print(f"\nFiles skipped: {len(skipped_files)}")
        for path in skipped_files:
            print(f"  ⚠️  {path.name}")

    print("\nNext steps:")
    print("1. Open the updated Network_Config files in Excel to verify the Products sheet")
    print("2. Adjust units_per_mix values if needed (get actual batch sizes from manufacturing)")
    print("3. Run tests to verify the parser can read the new column")
    print("4. Commit the updated files")


if __name__ == "__main__":
    main()
