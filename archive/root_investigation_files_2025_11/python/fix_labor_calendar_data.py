#!/usr/bin/env python3
"""Fix LaborCalendar data quality issues in Network_Config.xlsx."""

from openpyxl import load_workbook
import pandas as pd

def fix_labor_calendar():
    """Fill NaN values in LaborCalendar sheet with reasonable defaults."""

    file_path = 'data/examples/Network_Config.xlsx'

    # Read current data
    df = pd.read_excel(file_path, sheet_name='LaborCalendar', engine='openpyxl')

    print("Before fix:")
    print(f"  NaN in overtime_rate: {df['overtime_rate'].isnull().sum()}")
    print(f"  NaN in non_fixed_rate: {df['non_fixed_rate'].isnull().sum()}")

    # Fill NaN values with reasonable defaults
    # Overtime rate: typically 1.5× regular rate
    # Non-fixed rate: typically 2× regular rate (weekends/holidays)

    # Fill overtime_rate: use 1.5× regular_rate, or 30 if regular is also 0/NaN
    df['overtime_rate'] = df.apply(
        lambda row: row['overtime_rate'] if pd.notna(row['overtime_rate'])
                   else (row['regular_rate'] * 1.5 if row['regular_rate'] > 0 else 30.0),
        axis=1
    )

    # Fill non_fixed_rate: use 2× regular_rate, or 40 if regular is also 0/NaN
    df['non_fixed_rate'] = df.apply(
        lambda row: row['non_fixed_rate'] if pd.notna(row['non_fixed_rate'])
                   else (row['regular_rate'] * 2.0 if row['regular_rate'] > 0 else 40.0),
        axis=1
    )

    print("\nAfter fix:")
    print(f"  NaN in overtime_rate: {df['overtime_rate'].isnull().sum()}")
    print(f"  NaN in non_fixed_rate: {df['non_fixed_rate'].isnull().sum()}")
    print(f"  Sample overtime rates: {df['overtime_rate'].head(10).tolist()}")
    print(f"  Sample non_fixed rates: {df['non_fixed_rate'].head(10).tolist()}")

    # Write back to Excel
    wb = load_workbook(file_path)
    ws = wb['LaborCalendar']

    # Clear existing data (keep header)
    ws.delete_rows(2, ws.max_row)

    # Write updated data
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(file_path)
    print(f"\n✓ Updated {file_path}")
    print(f"  Fixed {len(df)} labor days")

if __name__ == "__main__":
    fix_labor_calendar()
