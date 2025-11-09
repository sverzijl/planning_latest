#!/usr/bin/env python3
"""Update Network_Config.xlsx with new cost parameters and improved formatting."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def update_cost_parameters():
    """Add freshness_incentive_weight and changeover_cost_per_start to CostParameters sheet."""

    file_path = 'data/examples/Network_Config.xlsx'
    wb = load_workbook(file_path)
    ws = wb['CostParameters']

    # Read existing data
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Skip empty rows
            data.append(list(row))

    # Define all parameters with explanations
    all_params = [
        # Core production/storage costs
        ('production_cost_per_unit', 1.30, '$/unit',
         'Base cost to manufacture one unit (materials, direct labor, overhead)'),

        # Storage costs - Unit-based (legacy)
        ('storage_cost_frozen_per_unit_day', 0.00, '$/(unit·day)',
         'Frozen storage per unit per day (use pallet-based instead)'),
        ('storage_cost_ambient_per_unit_day', 0.00, '$/(unit·day)',
         'Ambient storage per unit per day (use pallet-based instead)'),

        # Storage costs - Pallet-based (recommended)
        ('storage_cost_fixed_per_pallet', 0.00, '$/pallet',
         'Legacy fixed pallet entry cost (use state-specific instead)'),
        ('storage_cost_per_pallet_day_frozen', 0.98, '$/pallet/day',
         'Daily frozen storage cost per pallet (320 units)'),
        ('storage_cost_per_pallet_day_ambient', 0.00, '$/pallet/day',
         'Daily ambient storage cost per pallet (320 units)'),
        ('storage_cost_fixed_per_pallet_frozen', 14.26, '$/pallet',
         'Fixed cost when pallet enters frozen storage'),
        ('storage_cost_fixed_per_pallet_ambient', 0.00, '$/pallet',
         'Fixed cost when pallet enters ambient storage'),

        # Waste and shortage
        ('waste_cost_multiplier', 1.50, '-',
         'Multiplier on production cost for expired/wasted units'),
        ('shortage_penalty_per_unit', 10.00, '$/unit',
         'High penalty for unmet demand (encourages feasibility)'),

        # NEW: Freshness incentive (2025-10-22)
        ('freshness_incentive_weight', 0.05, '$/unit/age_ratio',
         'Staleness penalty scaled by age ratio (0-1). Encourages FIFO. 0.05 = mild preference'),

        # NEW: Changeover cost (2025-10-22)
        ('changeover_cost_per_start', 50.00, '$/changeover',
         'Cost per product startup (0→1 transition). Encourages campaign production. 50-200 typical'),
    ]

    # Create new dataframe
    df_new = pd.DataFrame(all_params, columns=['cost_type', 'value', 'unit', 'description'])

    # Clear existing sheet (keep header)
    ws.delete_rows(2, ws.max_row)

    # Write new data
    for r_idx, row_data in enumerate(dataframe_to_rows(df_new, index=False, header=False), start=2):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Apply formatting
            if c_idx == 1:  # cost_type column
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif c_idx == 2:  # value column
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif c_idx == 3:  # unit column
                cell.font = Font(italic=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif c_idx == 4:  # description column
                cell.font = Font(size=9, color='666666')
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            bottom=Side(style='medium', color='000000')
        )

    # Highlight NEW parameters
    new_param_rows = [r_idx for r_idx, row in enumerate(all_params, start=2)
                      if 'freshness_incentive' in row[0] or 'changeover_cost' in row[0]]

    for row_idx in new_param_rows:
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            if col_idx == 1:
                cell.font = Font(bold=True, size=10, color='2E7D32')

    # Adjust column widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 65

    # Set row heights
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 30

    ws.row_dimensions[1].height = 25

    # Save
    wb.save(file_path)
    print(f"✓ Updated {file_path}")
    print(f"  Added parameters:")
    print(f"    - freshness_incentive_weight: 0.05 $/unit/age_ratio")
    print(f"    - changeover_cost_per_start: 50.00 $/changeover")
    print(f"  Total parameters: {len(all_params)}")
    print(f"  NEW parameters highlighted in green")

if __name__ == "__main__":
    update_cost_parameters()
